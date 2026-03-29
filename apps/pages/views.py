import json
import gpxpy

import zipfile
import shapefile
import tempfile
import os

from django.shortcuts import render, redirect, get_object_or_404
from .models import *
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from apps.common.models import Farm, Parcel, CropPlan, CropType, Substance, ParcelAction, Tag, Invitation, FieldType, ParcelPolygon, SheetFile, SheetChat
from django.contrib.auth.decorators import login_required
from django.conf import settings
from django.http import JsonResponse
from django.urls import reverse
from apps.users.decorators import role_required
from apps.users.models import UserRole
from django.contrib import messages
from django.core.mail import send_mail
from django.contrib.auth import get_user_model

User = get_user_model()


def landing(request):
  return render(request, 'pages/landing.html')

@login_required(login_url='/users/signin/')
def dashboard(request):
  context = {
    'segment': 'dashboard',
    'title': 'Dashboard'
  }
  return render(request, "dashboard/index.html", context)

@login_required(login_url='/users/signin/')
def farms(request):
  farms = Farm.objects.filter(created_by=request.user)

  context = {
    'segment': 'farms',
    'farms': farms,
    'title': 'Farms'
  }

  return render(request, "pages/farms/index.html", context)

@login_required(login_url='/users/signin/')
def create_farm(request):
  tags = Tag.objects.all()

  if request.method == 'POST':
    name = request.POST.get('name')
    address = request.POST.get('address')
    tags = request.POST.getlist('tags')
    lat = request.POST.get("latitude")
    lon = request.POST.get("longitude")

    farm = Farm.objects.create(
      name=name,
      address=address,
      lat=lat,
      lon=lon,
      created_by=request.user
    )
    farm.tags.set(tags)

    Role.objects.create(
      user=request.user,
      farm=farm,
      role=UserRole.FARMER,
      active=True
    )

    request.user.active_farm = farm
    request.user.save()
    
    return redirect(reverse('farms'))

  context = {
    'segment': 'farm',
    'title': 'Create new Farm',
    'tags': tags,
    'API_KEY': getattr(settings, 'GOOGLE_MAP_API_KEY'),
  }
  return render(request, "pages/farms/new.html", context)

@login_required(login_url='/users/signin/')
def edit_farm(request, pk):
  farm = get_object_or_404(Farm, pk=pk)
  tags = Tag.objects.all()

  if request.method == 'POST':
    tags = request.POST.getlist('tags')

    farm.name = request.POST.get('name')
    farm.lat = request.POST.get("latitude")
    farm.lon = request.POST.get("longitude")
    farm.address = request.POST.get("address")

    farm.save()

    farm.tags.set(tags)

    return redirect(request.META.get('HTTP_REFERER'))
  
  context = {
    'farm': farm,
    'segment': 'farm',
    'title': 'Edit Farm',
    'tags': tags,
    'API_KEY': getattr(settings, 'GOOGLE_MAP_API_KEY'),
  }
  return render(request, "pages/farms/edit.html", context)

@login_required(login_url='/users/signin/')
def delete_farm(request, pk):
  farm = get_object_or_404(Farm, pk=pk)
  farm.delete()
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def farm_details(request, pk):
  farm = get_object_or_404(Farm, pk=pk)
  parcels = Parcel.objects.filter(farm=farm)

  context = {
    'farm': farm,
    'parcels': parcels,
    'title': farm.name,
  }
  return render(request, "pages/parcels/index.html", context)

def create_parcel(request, farm_id):
  farm = get_object_or_404(Farm, pk=farm_id)
  if request.method == 'POST':
    name = request.POST.get('name')
    info = request.POST.get('info')
    culture = request.POST.get('culture')

    Parcel.objects.create(
      farm=farm,
      name=name,
      info=info,
      culture=culture
    )

    return redirect(reverse('farm_details', args=[farm.id]))

  context = {
    'segment': 'farms'
  }
  return render(request, "pages/parcels/create.html", context)

def edit_parcel(request, pk):
  parcel = get_object_or_404(Parcel, pk=pk)
  if request.method == 'POST':
    parcel.name = request.POST.get('name')
    parcel.info = request.POST.get('info')
    parcel.culture = request.POST.get('culture')

    parcel.save()
    return redirect(reverse('farm_details', args=[parcel.farm.id]))

  context = {
    'segment': 'farms',
    'parcel': parcel
  }
  return render(request, "pages/parcels/edit.html", context)

def delete_parcel(request, pk):
  parcel = get_object_or_404(Parcel, pk=pk)
  parcel.delete()

  return redirect(request.META.get('HTTP_REFERER'))

import time
from pyproj import Transformer, CRS

def simplify_coords(coords, max_points=500):
  if len(coords) <= max_points:
    return coords
  step = max(1, len(coords) // max_points)
  return coords[::step]

@login_required(login_url='/users/signin/')
def import_parcel_polygon(request, pk):
  parcel = get_object_or_404(Parcel, pk=pk)

  if request.method == "POST":
    uploaded_file = request.FILES['file']
    filename = uploaded_file.name.lower()
    polygons_to_create = []

    if filename.endswith(".gpx"):
      gpx = gpxpy.parse(uploaded_file)

      for track in gpx.tracks:
        for segment in track.segments:
          if len(segment.points) > 20000:
            continue

          coords = [[p.latitude, p.longitude] for p in segment.points]

          if coords:
            coords = simplify_coords(coords)
            polygons_to_create.append(
              ParcelPolygon(parcel=parcel, polygon=coords)
            )

    elif filename.endswith(".zip"):
      with tempfile.TemporaryDirectory() as tmpdir:
        zip_path = os.path.join(tmpdir, uploaded_file.name)

        with open(zip_path, 'wb+') as f:
          for chunk in uploaded_file.chunks():
            f.write(chunk)

        with zipfile.ZipFile(zip_path, 'r') as z:
          z.extractall(tmpdir)

        shp_path = shx_path = dbf_path = None

        for root, dirs, files in os.walk(tmpdir):
          for file in files:
            if file.lower().endswith(".shp"):
              base = os.path.splitext(file)[0]
              shp = os.path.join(root, base + ".shp")
              shx = os.path.join(root, base + ".shx")
              dbf = os.path.join(root, base + ".dbf")

              if os.path.exists(shp) and os.path.exists(shx) and os.path.exists(dbf):
                shp_path = shp
                shx_path = shx
                dbf_path = dbf
                break
          if shp_path:
            break

        if not shp_path:
          raise ValueError("Valid shapefile not found")

        prj_path = shp_path.replace(".shp", ".prj")

        source_crs = CRS.from_epsg(32646)  # default fallback

        if os.path.exists(prj_path):
          try:
            with open(prj_path) as f:
              prj_text = f.read().strip()

            if prj_text:
              source_crs = CRS.from_wkt(prj_text)
          except Exception:
            source_crs = CRS.from_epsg(32646)

        transformer = Transformer.from_crs(source_crs, "EPSG:4326", always_xy=True)

        with open(shp_path, "rb") as shp_f, \
             open(shx_path, "rb") as shx_f, \
             open(dbf_path, "rb") as dbf_f:

          sf = shapefile.Reader(shp=shp_f, shx=shx_f, dbf=dbf_f)

          for shape in sf.shapes():
            if len(shape.parts) > 1:
              for i in range(len(shape.parts)):
                start = shape.parts[i]
                end = shape.parts[i + 1] if i + 1 < len(shape.parts) else len(shape.points)
                part = shape.points[start:end]

                coords = []
                for x, y in part:
                  lon, lat = transformer.transform(x, y)
                  coords.append([lat, lon])

                if coords:
                  coords = simplify_coords(coords)
                  polygons_to_create.append(
                    ParcelPolygon(parcel=parcel, polygon=coords)
                  )
            else:
              coords = []
              for x, y in shape.points:
                lon, lat = transformer.transform(x, y)
                coords.append([lat, lon])

              if coords:
                coords = simplify_coords(coords)
                polygons_to_create.append(
                  ParcelPolygon(parcel=parcel, polygon=coords)
                )

    else:
      raise ValueError("Unsupported file format")

    if polygons_to_create:
      ParcelPolygon.objects.bulk_create(polygons_to_create)

  return redirect(request.META.get('HTTP_REFERER'))


def parcel_details(request, pk):
  parcel = get_object_or_404(Parcel, pk=pk)

  parcel_polygon_qs = ParcelPolygon.objects.filter(parcel=parcel)
  parcel_polygons = list(parcel_polygon_qs.values('id', 'polygon'))

  for p in parcel_polygons:
    if len(p['polygon']) > 1000:
      p['polygon'] = p['polygon'][::5]

  all_coords = []
  for parcel_polygon in parcel_polygons:
    all_coords.extend(parcel_polygon['polygon'])

  if all_coords:
    avg_lat = sum(coord[0] for coord in all_coords) / len(all_coords)
    avg_lng = sum(coord[1] for coord in all_coords) / len(all_coords)
  else:
    avg_lat, avg_lng = 51.30, 0.7

  crop_plans = CropPlan.objects.filter(parcel_polygon__parcel=parcel).select_related('parcel_polygon', 'crop_type')
  crop_types = CropType.objects.all()
  substances = Substance.objects.all()

  context = {
    'parcel': parcel,
    'parcel_polygons': parcel_polygons,
    'farm': parcel.farm,
    'crop_plans': crop_plans,
    'crop_types': crop_types,
    'substances': substances,
    'API_KEY': getattr(settings, 'GOOGLE_MAP_API_KEY'),
    'avg_lat': avg_lat,
    'avg_lng': avg_lng
  }

  return render(request, "pages/parcels/detail.html", context)

@login_required(login_url='/users/signin/')
def add_farm_manager(request, pk):
  farm = get_object_or_404(Farm, pk=pk)
  if request.method == 'POST':
    user_id = request.POST.get('user')
    user = get_object_or_404(User, pk=user_id)

    role, created = Role.objects.update_or_create(
      farm=farm,
      role=UserRole.FARMER,
      defaults={
        'user': user,
        'active': True
      }
    )

    return redirect(request.META.get('HTTP_REFERER'))
  
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def save_parcel_polygon(request, pk):
  parcel = get_object_or_404(Parcel, pk=pk)

  if request.method == "POST":
    polygon_data = request.POST.get("polygon")
    parcel_id = request.POST.get("parcel_id")

    if polygon_data:
      if parcel_id:
        parcel_polygon = get_object_or_404(ParcelPolygon, id=parcel_id, parcel=parcel)
        parcel_polygon.polygon = json.loads(polygon_data)
        parcel_polygon.save()
      else:
        ParcelPolygon.objects.create(
          parcel=parcel,
          polygon=json.loads(polygon_data)
        )

  return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url='/users/signin/')
def delete_parcel_polygon(request, parcel_id, parcel_polygon_id):
  parcel = get_object_or_404(Parcel, pk=parcel_id)
  parcel_polygon = get_object_or_404(ParcelPolygon, pk=parcel_polygon_id, parcel=parcel)

  parcel_polygon.delete()

  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def create_crop_plan(request, parcel_polygon_id):
  parcel_polygon = get_object_or_404(ParcelPolygon, id=parcel_polygon_id)

  if request.method == "POST":
    crop_type = request.POST.get("crop_type")
    year = request.POST.get("year")

    CropPlan.objects.create(
      parcel_polygon=parcel_polygon,
      crop_type_id=crop_type,
      year=year
    )

  return redirect(request.META.get("HTTP_REFERER"))

@login_required(login_url='/users/signin/')
def add_action(request, crop_plan_id):
  crop_plan = get_object_or_404(CropPlan, id=crop_plan_id)

  if request.method == "POST":
    action_type = request.POST.get("action_type")
    substance = request.POST.get("substance")
    date = request.POST.get("date")

    ParcelAction.objects.create(
      crop_plan=crop_plan,
      action_type=action_type,
      substance_id=substance if substance else None,
      date=date
    )

  return redirect(request.META.get("HTTP_REFERER"))

@login_required(login_url='/users/signin/')
def parcel_plans(request, parcel_polygon_id):
  plans = CropPlan.objects.filter(parcel_polygon_id=parcel_polygon_id).select_related("crop_type")
  data = []
  for p in plans:
    actions = []

    for a in p.actions.all():
      actions.append({
        "type": a.action_type,
        "date": str(a.date),
        "substance": a.substance.name if a.substance else None
      })

    data.append({
      "id": p.id,
      "crop": p.crop_type.name,
      "year": p.year,
      "actions": actions
    })

  return JsonResponse(data, safe=False)


#
from apps.common.models import Tab, Sheet, TabFields, TabRow, Asset, TabCell

@login_required(login_url='/users/signin/')
def tab_list(request):
  tabs = Tab.objects.all()
  sheets = Sheet.objects.all()

  page_number = request.GET.get('page', 1) 
  paginator = Paginator(tabs, 9)
  try:
    tabs = paginator.page(page_number)
  except PageNotAnInteger:
    tabs = paginator.page(1)
  except EmptyPage:
    tabs = paginator.page(paginator.num_pages)

  context = {
    'tabs': tabs,
    'sheets': sheets,
    'segment': 'tab',
    'title': 'Tabs'
  }
  return render(request, 'pages/tabs/index.html', context)

def add_new_field(request, pk):
  tab = get_object_or_404(Tab, pk=pk)

  if request.method == 'POST':
    name = request.POST.get('name')
    type = request.POST.get('type')

    TabFields.objects.update_or_create(
      tab=tab,
      name=name,
      defaults={'type': type}
    )

    return redirect(request.META.get('HTTP_REFERER'))

  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def edit_field(request, pk):
  field = get_object_or_404(TabFields, pk=pk)
  if request.method == 'POST':
    name = request.POST.get('name')
    type = request.POST.get('type')

    field.name = name
    field.type = type
    field.save()

    return redirect(request.META.get('HTTP_REFERER'))

  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def delete_field(request, pk):
  field = get_object_or_404(TabFields, pk=pk)
  field.delete()
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def add_data(request, pk):
  tab = get_object_or_404(Tab, pk=pk)
  fields = TabFields.objects.filter(tab=tab)
  rows = TabRow.objects.filter(tab=tab).prefetch_related("cells")

  if request.method == "POST":

    delete_rows = request.POST.getlist("delete_rows")

    rows_data = {}

    for key, value in request.POST.items():
      if key.startswith("data"):
        row_key, field_id = key.replace("data[", "").replace("]", "").split("[")

        if row_key not in rows_data:
          rows_data[row_key] = {}

        rows_data[row_key][field_id] = value

    for row_key, cells in rows_data.items():

      if row_key.startswith("row"):
        row_id = row_key.replace("row", "")

        # skip deleted rows
        if row_id in delete_rows:
          continue

        tab_row = TabRow.objects.get(id=row_id)

      else:
        last = TabRow.objects.filter(tab=tab).order_by("-row_index").first()
        next_index = 1 if not last else last.row_index + 1

        tab_row = TabRow.objects.create(
          tab=tab,
          row_index=next_index
        )

      for field_id, value in cells.items():
        TabCell.objects.update_or_create(
          row=tab_row,
          field_id=field_id,
          defaults={"value": value}
        )

    # finally delete rows
    if delete_rows:
      TabRow.objects.filter(id__in=delete_rows).delete()

    return redirect("add_data", pk=pk)

  context = {
    "tab": tab,
    "fields": fields,
    "rows": rows
  }

  return render(request, "pages/sheets/add_data.html", context)

@login_required(login_url='/users/signin/')
def create_tab(request, sheet_id):
  sheet = get_object_or_404(Sheet, pk=sheet_id)
  if request.method == 'POST':
    name = request.POST.get('name')
    Tab.objects.get_or_create(
      name=name,
      sheet=sheet
    )
    return redirect(request.META.get('HTTP_REFERER'))
  
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def edit_tab(request, pk):
  tab = get_object_or_404(Tab, pk=pk)
  if request.method == 'POST':
    name = request.POST.get('name')
    tab.name = name
    tab.save()
    return redirect(request.META.get('HTTP_REFERER'))
  
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def delete_tab(request, pk):
  tab = get_object_or_404(Tab, pk=pk)
  tab.delete()
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def tab_row_edit(request, pk):
  row = get_object_or_404(TabRow, pk=pk)
  tab = row.tab

  fields = TabFields.objects.filter(tab=tab)


  context = {
    "row": row,
    "fields": fields,
    "segment": "tab",
    "title": "Tab Rows"
  }

  return render(request, "pages/tabs/tab_row_edit.html", context)

@login_required(login_url='/users/signin/')
def tab_row_delete(request, pk):
  row = get_object_or_404(TabRow, pk=pk)
  row.delete()
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def tab_row_upload(request, pk):
  row = get_object_or_404(TabRow, pk=pk)

  if request.method == "POST":
    file = request.FILES.get("asset")

    if file:
      Asset.objects.update_or_create(
        user=request.user,
        row=row,
        defaults={
          'file': file
        }
      )

      return redirect(request.META.get('HTTP_REFERER'))
  
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def personnel(request):
  context = {
    'segment': 'personnel',
    'title': 'Personnel'
  }
  return render(request, 'pages/farms/personnel.html', context)

@login_required(login_url='/users/signin/')
def tasks(request):
  context = {
    'segment': 'tasks',
    'title': 'Tasks'
  }
  return render(request, 'pages/farms/tasks.html', context)

@login_required(login_url='/users/signin/')
def review_docs(request):
  context = {
    'segment': 'review_docs',
    'title': 'Review Docs'
  }
  return render(request, 'pages/farms/review_docs.html', context)

@login_required(login_url='/users/signin/')
def pre_audit(request):
  context = {
    'segment': 'pre_audit',
    'title': 'Pre Audit'
  }
  return render(request, 'pages/farms/pre_audit.html', context)

@login_required(login_url='/users/signin/')
def search(request):
  context = {
    'segment': 'search',
    'title': 'Search'
  }
  return render(request, 'pages/farms/search.html', context)


###
from apps.common.forms import SheetForm

@login_required(login_url='/users/signin/')
def certification(request):
  sheets = Sheet.objects.filter(farm=request.user.active_farm)

  if request.method == 'POST':
    name = request.POST.get('name')
    Sheet.objects.get_or_create(
      name=name
    )
    return redirect(request.META.get('HTTP_REFERER'))
  
  context = {
    'segment': 'certification',
    'title': 'Certification',
    'sheets': sheets
  }
  return render(request, 'pages/sheets/index.html', context)

@login_required(login_url='/users/signin/')
def create_sheet(request):
  form = SheetForm()
  if request.method == 'POST':
    form = SheetForm(request.POST, request.FILES)
    if form.is_valid():
      sheet = form.save(commit=False)
      sheet.farm = request.user.active_farm
      sheet.save()

      return redirect(request.META.get('HTTP_REFERER'))

  context = {
    'form': form
  }
  return render(request, 'pages/sheets/create.html', context)

@login_required(login_url='/users/signin/')
def edit_sheet(request, pk):
  sheet = get_object_or_404(Sheet, pk=pk)
  form = SheetForm(instance=sheet, user=request.user)
  if request.method == 'POST':
    form = SheetForm(request.POST, request.FILES, instance=sheet, user=request.user)
    if form.is_valid():
      form.save()

      return redirect(request.META.get('HTTP_REFERER'))
  
  context = {
    'form': form,
    'sheet': sheet
  }
  return render(request, 'pages/sheets/edit.html', context)

@login_required(login_url='/users/signin/')
def delete_sheet(request, pk):
  sheet = get_object_or_404(Sheet, pk=pk)
  sheet.delete()
  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def download_sheet_file(request, pk):
  sheet = get_object_or_404(Sheet, id=pk)
  file_path = sheet.file.path

  if not os.path.exists(file_path):
    raise Http404("File not found")

  response = FileResponse(open(file_path, "rb"), as_attachment=True)
  response["Content-Disposition"] = f'attachment; filename="{sheet.filename}"'

  return response


def sheet_details(request, pk):
  sheet = get_object_or_404(Sheet, pk=pk)
  tabs = Tab.objects.filter(sheet=sheet)
  form = SheetForm(instance=sheet, user=request.user)

  if request.method == 'POST':
    form = SheetForm(request.POST, request.FILES, instance=sheet, user=request.user)
    if form.is_valid():
      form.save()

      return redirect(request.META.get('HTTP_REFERER'))

  context = {
    'sheet': sheet,
    'tabs': tabs,
    'form': form
  }
  return render(request, 'pages/sheets/sheet-info.html', context)

def sheet_media(request, pk):
  sheet = get_object_or_404(Sheet, pk=pk)
  tabs = Tab.objects.filter(sheet=sheet)
  files = SheetFile.objects.filter(sheet=sheet)

  context = {
    'sheet': sheet,
    'tabs': tabs,
    'files': files
  }
  return render(request, 'pages/sheets/sheet-media.html', context)

def sheet_chat(request, pk):
  sheet = get_object_or_404(Sheet, pk=pk)
  tabs = Tab.objects.filter(sheet=sheet)
  files = SheetFile.objects.filter(sheet=sheet)
  chats = SheetChat.objects.filter(sheet=sheet).select_related('sender').order_by('created_at')

  if request.method == 'POST':
    message = request.POST.get('message')
    if message:
      SheetChat.objects.create(
        sheet=sheet,
        sender=request.user,
        message=message
      )

    return redirect(request.META.get('HTTP_REFERER'))

  context = {
    'sheet': sheet,
    'tabs': tabs,
    'files': files,
    'chats': chats
  }
  return render(request, 'pages/sheets/sheet-chat.html', context)

def upload_sheet_file(request, pk):
  sheet = get_object_or_404(Sheet, pk=pk)
  if request.method == 'POST':
    file = request.FILES.get('file')

    SheetFile.objects.create(
      file=file,
      sheet=sheet
    )

    return redirect(request.META.get('HTTP_REFERER'))

  return redirect(request.META.get('HTTP_REFERER'))

def download_sheet_file(request, pk):
  sheet_file = get_object_or_404(SheetFile, pk=pk)
  file_path = sheet_file.file.path

  if not os.path.exists(file_path):
    raise Http404("File not found")

  response = FileResponse(open(file_path, "rb"), as_attachment=True)
  response["Content-Disposition"] = f'attachment; filename="{sheet_file.filename}"'

  return response

def delete_sheet_file(request, pk):
  sheet_file = get_object_or_404(SheetFile, pk=pk)
  sheet_file.delete()
  return redirect(request.META.get('HTTP_REFERER'))

def tab_details(request, pk):
  tab = get_object_or_404(Tab, pk=pk)
  tabs = Tab.objects.filter(sheet=tab.sheet)

  fields = TabFields.objects.filter(tab=tab)
  rows = TabRow.objects.filter(tab=tab).prefetch_related("cells")

  if request.method == "POST":
    delete_rows = request.POST.getlist("delete_rows")

    rows_data = {}
    for key, value in request.POST.items():
      if key.startswith("data"):
        row_key, field_id = key.replace("data[", "").replace("]", "").split("[")
        if row_key not in rows_data:
          rows_data[row_key] = {}
        rows_data[row_key][field_id] = value

    for key, file in request.FILES.items():
      if key.startswith("file[row"):
        row_id = key.replace("file[row", "").replace("]", "")
        try:
          row = TabRow.objects.get(id=row_id)
          Asset.objects.update_or_create(
            row=row,
            defaults={
              "file": file,
              "user": request.user
            }
          )
        except TabRow.DoesNotExist:
          pass

    for row_key, cells in rows_data.items():
      if row_key.startswith("row"):
        row_id = row_key.replace("row", "")
        if row_id in delete_rows:
          continue
        tab_row = TabRow.objects.get(id=row_id)
      else:
        last = TabRow.objects.filter(tab=tab).order_by("-row_index").first()
        next_index = 1 if not last else last.row_index + 1
        tab_row = TabRow.objects.create(tab=tab, row_index=next_index)

      for field_id, value in cells.items():
        TabCell.objects.update_or_create(
          row=tab_row,
          field_id=field_id,
          defaults={"value": value}
        )

    field_names = request.POST.getlist("field_names")
    for field_id, value in request.POST.items():
      if field_id.startswith("field_names["):
        fid = field_id.replace("field_names[", "").replace("]", "")
        TabFields.objects.filter(id=fid, tab=tab).update(name=value)

    if delete_rows:
      TabRow.objects.filter(id__in=delete_rows).delete()

    return redirect(request.META.get('HTTP_REFERER'))

  context = {
    'segment': 'certification',
    'active_tab': tab,
    'tabs': tabs,
    'tab': tab,
    'sheet': tab.sheet,
    'fields': fields,
    'rows': rows,
    'types': FieldType.choices
  }
  return render(request, 'pages/sheets/details.html', context)

def row_files(request, pk):
  row = get_object_or_404(TabRow, pk=pk)
  files = Asset.objects.filter(user=request.user, row=row)

  context = {
    'segment': 'certification',
    'row': row,
    'files': files
  }
  return render(request, 'pages/sheets/files.html', context)


def upload_file(request, pk):
  if request.method != "POST":
    return JsonResponse({"error": "Invalid request"}, status=400)

  if not request.user.is_authenticated:
    return JsonResponse({"error": "Unauthorized"}, status=401)

  row = get_object_or_404(TabRow, pk=pk)

  file = request.FILES.get('file')
  if not file:
    return JsonResponse({"error": "No file provided"}, status=400)

  asset = Asset.objects.create(
    user=request.user,
    file=file,
    row=row
  )

  return JsonResponse({
    "status": "success",
    "id": asset.id,
    "filename": asset.file.name,
    "url": asset.file.url,
  })

def delete_file(request, pk):
  if request.method == "POST":
    asset = get_object_or_404(Asset, pk=pk, user=request.user)
    asset.delete()
    return JsonResponse({"status": "deleted"})

  return JsonResponse({"error": "Invalid request"}, status=400)

import csv, os
from django.http import HttpResponse, FileResponse, Http404

def download_asset(request, row_id):
  try:
    asset = Asset.objects.get(pk=row_id)
  except Asset.DoesNotExist:
    raise Http404("File not found")

  file_path = asset.file.path

  if not os.path.exists(file_path):
    raise Http404("File not found")

  response = FileResponse(open(file_path, "rb"), as_attachment=True)
  response["Content-Disposition"] = f'attachment; filename="{asset.filename}"'

  return response

def download_tab_csv(request, pk):
  tab = get_object_or_404(Tab, pk=pk)

  fields = TabFields.objects.filter(tab=tab).order_by("id")
  rows = TabRow.objects.filter(tab=tab).prefetch_related("cells").order_by("row_index")

  response = HttpResponse(content_type="text/csv")
  response["Content-Disposition"] = f'attachment; filename="{tab.name}.csv"'

  writer = csv.writer(response)

  writer.writerow([f.name for f in fields])

  for row in rows:
    cell_map = {c.field_id: c.value for c in row.cells.all()}
    writer.writerow([cell_map.get(f.id, "") for f in fields])

  return response

####
@login_required(login_url='/users/signin/')
def reports(request):
  context = {
    'segment': 'reports',
    'title': 'Reports'
  }
  return render(request, 'pages/farms/reports.html', context)


# Request
from apps.common.models import Role

@login_required(login_url='/users/signin/')
def role_request(request):
  pending_requests = Role.objects.filter(
    farm=request.user.active_farm,
    active=False
  )
  context = {
    'segment': 'request',
    'parent': 'personnel',
    'title': 'Role Request',
    'pending_requests': pending_requests
  }
  return render(request, 'pages/farms/role_request.html', context)

@login_required(login_url='/users/signin/')
def onboarded_roles(request):
  onboarded_roles = Role.objects.filter(
    farm=request.user.active_farm,
    active=True
  )
  context = {
    'segment': 'onboarded',
    'parent': 'personnel',
    'title': 'Onboarded Roles',
    'onboarded_roles': onboarded_roles
  }
  return render(request, 'pages/farms/onboarded.html', context)

@login_required(login_url='/users/signin/')
def pending_invitations(request):
  pending_invitations = Invitation.objects.filter(
    farm=request.user.active_farm,
    accepted=False
  )
  context = {
    'segment': 'pending',
    'parent': 'personnel',
    'title': 'Pending Roles',
    'pending_invitations': pending_invitations
  }
  return render(request, 'pages/farms/pending.html', context)


@login_required(login_url='/users/signin/')
def remove_invitation(request, pk):
  invitation = get_object_or_404(Invitation, pk=pk)
  invitation.delete()

  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def accept_request(request, pk):
  role = get_object_or_404(Role, pk=pk)
  role.active = True
  role.save()

  role_user = role.user
  role_user.active_farm = role.farm
  role_user.save()

  return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def reject_request(request, pk):
  role = get_object_or_404(Role, pk=pk)
  role.delete()
  return redirect(request.META.get('HTTP_REFERER'))


def send_invitation_email(invitation):
  existing_user = User.objects.filter(email=invitation.email).exists()

  if existing_user:
    invite_link = f"{settings.SITE_URL}/invitations/accept/?invite_token={invitation.token}"
  else:
    invite_link = f"{settings.SITE_URL}/users/signup/?invite_token={invitation.token}"

  subject = f"You have been invited to join {invitation.farm.name}"

  message = f"""
You have been invited to join the farm "{invitation.farm.name}" with the role "{invitation.role}".

Accept the invitation using the link below:
{invite_link}
"""

  send_mail(
    subject,
    message,
    settings.DEFAULT_FROM_EMAIL,
    [invitation.email],
    fail_silently=False
  )

@login_required(login_url='/users/signin/')
def invite_personnel(request):
  roles = [
    {'value': v, 'label': l}
    for v, l in UserRole.choices
    if v not in ['ADMIN', 'FARMER']
  ]

  if request.method == 'POST':
    farm_id = request.POST.get('farm')
    role = request.POST.get('role')
    user_id = request.POST.get('user')
    email = request.POST.get('email')

    farm = get_object_or_404(Farm, pk=farm_id)

    if user_id:
      user = get_object_or_404(User, pk=user_id)
      email = user.email

    invitation, created = Invitation.objects.get_or_create(
      email=email,
      farm=farm,
      role=role,
      accepted=False
    )

    if created:
      send_invitation_email(invitation)

  context = {
    'segment': 'invite',
    'parent': 'personnel',
    'title': 'Invite Personnel',
    'roles': roles,
    'users': User.objects.all()
  }

  return render(request, 'pages/farms/invite_personnel.html', context)

@login_required(login_url='/users/signin/')
def accept_invitation(request):
  token = request.GET.get('invite_token')

  invitation = get_object_or_404(
    Invitation,
    token=token,
    accepted=False
  )

  if request.user.email != invitation.email:
    messages.error(request, "This invitation is not for your account.")
    return redirect("dashboard")

  Role.objects.update_or_create(
    farm=invitation.farm,
    user=request.user,
    role=invitation.role,
    defaults={
      'active': True
    }
  )

  request.user.active_farm = invitation.farm
  request.user.save()

  invitation.accepted = True
  invitation.save()

  messages.success(request, "Invitation accepted successfully.")
  return redirect(reverse('settings'))


@login_required(login_url='/users/signin/')
def farms_to_request(request):
  roles = [
    {'value': v, 'label': l}
    for v, l in UserRole.choices
    if v not in ['ADMIN', 'FARMER']
  ]

  farms = Farm.objects.exclude(
    farm_role__user=request.user,
    farm_role__active=True
  ).filter(
    farm_role__role='FARMER',
    farm_role__active=True
  ).distinct()

  context = {
    'segment': 'farms',
    'title': 'Farms to Request',
    'roles': roles,
    'farms': farms
  }
  return render(request, 'pages/farms/farms_to_request.html', context)


@login_required(login_url='/users/signin/')
def send_request(request, pk):
  farm = get_object_or_404(Farm, pk=pk)
  if request.method == 'POST':
    role = request.POST.get('role')

    Role.objects.get_or_create(
      farm=farm,
      user=request.user,
      role=role,
      active=False
    )

    request.user.active_farm = farm
    request.user.save()

    return redirect(request.META.get('HTTP_REFERER'))

  return redirect(request.META.get('HTTP_REFERER'))


# Preivew
import openpyxl
import xlrd
from zipfile import BadZipFile

def upload_xlsx_preview(request):
  if request.method == "POST" and request.FILES.get("file"):
    file = request.FILES["file"]
    filename = file.name.lower()

    preview_data = {}

    try:
      if filename.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file, data_only=True)

        for sheet_name in wb.sheetnames:
          ws = wb[sheet_name]
          rows = list(ws.iter_rows(values_only=True))

          if not rows:
            continue

          headers = [str(h) for h in rows[0]]

          preview_data[sheet_name] = {
            "fields": headers,
            "rows": [
              [str(cell) if cell is not None else "" for cell in row]
              for row in rows[1:]
            ]
          }

      elif filename.endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file.read())

        for sheet in wb.sheets():
          rows = []

          for row_idx in range(sheet.nrows):
            rows.append(sheet.row_values(row_idx))

          if not rows:
            continue

          headers = [str(h) for h in rows[0]]

          preview_data[sheet.name] = {
            "fields": headers,
            "rows": [
              [str(cell) if cell is not None else "" for cell in row]
              for row in rows[1:]
            ]
          }

      else:
        messages.error(request, "Only .xls and .xlsx files are supported.")
        return redirect(request.META.get('HTTP_REFERER'))

    except BadZipFile:
      messages.error(request, "Invalid .xlsx file.")
      return redirect(request.META.get('HTTP_REFERER'))

    except Exception as e:
      messages.error(request, f"Error processing file: {str(e)}")
      return redirect(request.META.get('HTTP_REFERER'))

    if not preview_data:
      messages.error(request, "No data found in file.")
      return redirect(request.META.get('HTTP_REFERER'))

    request.session["xlsx_preview"] = preview_data
    request.session["active_sheet"] = list(preview_data.keys())[0]
    request.session["xlsx_file_name"] = file.name.rsplit(".", 1)[0]

    return redirect(reverse("xlsx_preview_page"))

  return redirect(request.META.get('HTTP_REFERER'))

def xlsx_preview_page(request):
  preview = request.session.get("xlsx_preview")
  active_sheet = request.GET.get("sheet") or request.session.get("active_sheet")

  if not preview:
    return redirect(request.META.get('HTTP_REFERER'))

  if active_sheet not in preview:
    active_sheet = list(preview.keys())[0]

  request.session["active_sheet"] = active_sheet

  context = {
    "tabs": preview.keys(),
    "active_tab": active_sheet,
    "fields": preview[active_sheet]["fields"],
    "rows": preview[active_sheet]["rows"],
  }

  return render(request, "pages/sheets/preview.html", context)

def save_xlsx_to_db(request):
  preview_data = request.session.get("xlsx_preview")
  file_name = request.session.get("xlsx_file_name")

  if not preview_data:
    return redirect(request.META.get('HTTP_REFERER'))

  field_names = {}
  for key, value in request.POST.items():
    if key.startswith("field_names["):
      index = int(key[len("field_names["):-1])
      field_names[index] = value.strip()

  fields = [field_names[i] for i in sorted(field_names.keys())]

  rows_dict = {}
  for key, value in request.POST.items():
    if key.startswith("data["):
      parts = key[len("data["):-1].split("][")
      row_key, col_index = parts[0], int(parts[1])
      if row_key not in rows_dict:
        rows_dict[row_key] = {}

      rows_dict[row_key][col_index] = value

  rows = []
  for row_key in rows_dict:
    col_map = rows_dict[row_key]
    row = [col_map.get(i, "") for i in range(len(fields))]
    rows.append(row)

  active_sheet = request.session.get("active_sheet")
  sheet = Sheet.objects.create(name=file_name)

  for sheet_name, data in preview_data.items():
    if sheet_name == active_sheet:
      sheet_fields = fields
      sheet_rows = rows
    else:
      sheet_fields = data["fields"]
      sheet_rows = data["rows"]

    tab = Tab.objects.create(name=sheet_name, sheet=sheet)

    field_map = {}
    for field_name in sheet_fields:
      field = TabFields.objects.create(tab=tab, name=field_name, type="STRING")
      field_map[field_name] = field

    for row_index, row_values in enumerate(sheet_rows, start=1):
      tab_row = TabRow.objects.create(tab=tab, row_index=row_index)
      for col_name, value in zip(sheet_fields, row_values):
        TabCell.objects.create(row=tab_row, field=field_map[col_name], value=value)

  del request.session["xlsx_preview"]
  del request.session["xlsx_file_name"]

  first_tab = Tab.objects.filter(sheet=sheet).first()
  return redirect("tab_details", pk=first_tab.id)

def update_xlsx_session_sheet(request):
  if request.method == "POST":
    preview_data = request.session.get("xlsx_preview")
    active_sheet = request.session.get("active_sheet")

    if not preview_data or not active_sheet:
        return JsonResponse({"status": "error"}, status=400)

    body = json.loads(request.body)
    fields = body.get("fields", [])
    rows = body.get("rows", [])

    preview_data[active_sheet]["fields"] = fields
    preview_data[active_sheet]["rows"] = rows

    request.session["xlsx_preview"] = preview_data
    request.session.modified = True

    return JsonResponse({"status": "ok"})

  return JsonResponse({"status": "error"}, status=405)